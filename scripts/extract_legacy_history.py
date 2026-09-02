"""One-off extraction of the user's JSA reference workbook (corn/soybean
settlement history back to 2008) into a flat CSV bundled with the app.

Source: "Futures History.xlsx" -- one sheet per contract month (e.g. "CZ" =
corn December, "SX" = soybean November), with a "History" date column plus
one column per contract-year (e.g. "ZCZ08".."ZCZ21"), each holding that
contract's last ~300 trading sessions before expiration. No wheat sheets
exist in this file, so only ZC/ZS are extracted.

Re-run this only if the user supplies an updated version of the source
workbook -- it's a static backfill, not a live data feed.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

SOURCE_PATH = (
    r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents"
    r"\Research Analyst\Misc\Future Seasonal Charts\Futures History.xlsx"
)
OUT_PATH = Path(__file__).parent.parent / "legacy_futures_history.csv"

SHEET_TO_PRODUCT = {
    "SX": "ZS", "SF": "ZS", "SH": "ZS", "SK": "ZS", "SN": "ZS", "SQ": "ZS", "SU": "ZS",
    "CZ": "ZC", "CH": "ZC", "CK": "ZC", "CN": "ZC", "CU": "ZC",
}


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        product_code = SHEET_TO_PRODUCT.get(sheet_name)
        if not product_code:
            print(f"Skipping unrecognized sheet {sheet_name!r}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        header = data[0]
        for row in data[1:]:
            day = row[0]
            if day is None:
                continue
            for col_idx in range(1, len(header)):
                ticker = header[col_idx]
                price = row[col_idx]
                if ticker is None or price is None:
                    continue
                month_letter = ticker[2]
                year = 2000 + int(ticker[3:5])
                rows.append({
                    "product_code": product_code,
                    "month_letter": month_letter,
                    "contract_year": year,
                    "date": day.date().isoformat(),
                    "settle": float(price),
                })

    df = pd.DataFrame(rows).sort_values(["product_code", "month_letter", "contract_year", "date"])
    df.to_csv(OUT_PATH, index=False)
    print(f"Wrote {len(df):,} rows to {OUT_PATH}")
    print(df.groupby(["product_code", "month_letter"])["contract_year"].agg(["min", "max", "nunique"]))


if __name__ == "__main__":
    main()
